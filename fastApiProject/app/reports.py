import pymysql
from decimal import Decimal
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries, get_column_letter
import datetime
from cn2an import an2cn
from typing import List, Dict, Optional, Union
from collections import defaultdict
from pathlib import Path
import fnmatch
import logging
from .config import settings
from .utils import DatabaseManager

logger = logging.getLogger(__name__)


class TaxReportGenerator:
    # 默认值配置
    DEFAULT_COMPANY = "云策企服（驻马店）人力资源服务有限公司"
    DEFAULT_CREDIT_CODE = "91411700MAEFDQ7P7T"
    # 模板字体配置（与Excel模板保持一致）
    DEFAULT_FONT = Font(name='宋体', size=10)
    DATA_FONT = Font(name='宋体', size=9)
    BOLD_FONT = Font(name='宋体', size=10, bold=True)

    def __init__(self, db_config: dict):
        self.db_config = db_config
        # 初始化模板路径
        self.template_path = self._get_template_path()

    def _get_template_path(self) -> Path:
        """获取模板文件的跨平台路径"""
        app_dir = Path(__file__).parent
        template_path = app_dir / "templates" / "tax_template.xlsx"

        if not template_path.exists():
            raise FileNotFoundError(
                f"Excel模板文件未找到，请确保文件存在于: {template_path}\n"
                "解决方案：\n"
                "1. 在项目目录中创建 `app/templates/` 文件夹\n"
                f"2. 将您的模板文件复制到该文件夹并重命名为 `tax_template.xlsx`"
            )
        return template_path

    def query_tax_data(self, year_month: str, enterprise_ids: Optional[Union[int, List[int], str]] = None,
                       amount_type: int = 1) -> List[Dict]:
        """
        查询税务数据。
        - 如果提供了enterprise_ids (int)，则查询特定企业的数据。
        - 如果提供了enterprise_ids (list or str)，则查询指定ID列表的企业数据。
        - 否则，查询指定月份所有企业的数据。

        amount_type: 1=pay_amount(含服务费), 2=worker_pay_amount(不含服务费), 3=bill_amount(账单金额)
        """
        allowed_amount_fields = {
            1: 'pay_amount',
            2: 'worker_pay_amount',
            3: 'bill_amount'
        }
        amount_field = allowed_amount_fields.get(amount_type)
        if not amount_field:
            raise ValueError(f"无效的 amount_type: {amount_type}")

        try:
            with DatabaseManager(self.db_config) as conn:
                with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 【优化】将默认值直接在SQL中赋予，简化后续处理
                    sql = f"""
                    SELECT
                        realname AS 纳税人姓名,
                        credential_num AS 身份证号,
                        ROUND({amount_field}, 2) AS 营业额_元,
                        ROUND(tax_amount, 2) AS tax_amount,
                        enterprise_name,
                        enterprise_id,
                        tax_id AS 税地ID,
                        tax_address AS 税地名称,
                        service_pay_status,
                        tax_pay_status,
                        0.00 AS 增值税_元,
                        0.00 AS 教育费附加_元,
                        0.00 AS 地方教育附加_元,
                        0.00 AS 城市维护建设费_元,
                        0.00 AS 个人经营所得税税率,
                        0.00 AS 应纳个人经营所得税_元,
                        enterprise_name AS 备注
                    FROM
                        biz_balance_worker
                    WHERE
                        pay_status = 3
                        AND enterprise_id <> 36
                        AND DATE_FORMAT(payment_over_time, '%%Y-%%m') = %s
                    """
                    params = [year_month]

                    if enterprise_ids is not None:
                        ids_to_query = []
                        if isinstance(enterprise_ids, int):
                            ids_to_query = [enterprise_ids]
                        elif isinstance(enterprise_ids, str):
                            ids_to_query = [int(eid.strip()) for eid in enterprise_ids.split(',') if eid.strip()]
                        elif isinstance(enterprise_ids, list):
                            ids_to_query = enterprise_ids

                        if ids_to_query:
                            placeholders = ', '.join(['%s'] * len(ids_to_query))
                            sql += f" AND enterprise_id IN ({placeholders})"
                            params.extend(ids_to_query)

                    sql += " ORDER BY enterprise_name, id"
                    cursor.execute(sql, tuple(params))
                    return cursor.fetchall()
        except Exception as e:
            logger.error(f"数据库查询错误: {str(e)}")
            return []

    def is_merged(self, ws: Worksheet, merge_range: str) -> bool:
        """检查单元格范围是否已合并"""
        return merge_range in [str(r) for r in ws.merged_cells.ranges]

    def safe_write_merged_cell(self, ws: Worksheet, start_cell: str, value, merge_range: Optional[str] = None,
                               bold: bool = False, align: str = 'center', font: Optional[Font] = None):
        """安全写入合并单元格，并可设置对齐方式和字体"""
        alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell_font = font if font else (self.BOLD_FONT if bold else self.DEFAULT_FONT)

        if merge_range:
            if self.is_merged(ws, merge_range):
                ws.unmerge_cells(merge_range)
            ws.merge_cells(merge_range)
            first_cell = ws[start_cell]
            first_cell.value = value
            first_cell.font = cell_font
            first_cell.alignment = alignment
            for row in ws[merge_range]:
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = alignment
        else:
            cell = ws[start_cell]
            cell.value = value
            cell.font = cell_font
            cell.alignment = alignment

    def amount_to_cn(self, amount: float) -> str:
        """将金额转换为中文大写"""
        try:
            return an2cn(f"{amount:.2f}", "rmb")
        except Exception as e:
            logger.error(f"金额转换错误: {str(e)}")
            return "金额转换错误"

    def format_currency(self, amount: float) -> str:
        """格式化金额为￥XXX,XXX.XX格式"""
        return f"￥{amount:,.2f}"

    def add_border(self, ws: Worksheet, start_row: int, end_row: int, cols: List[str]):
        """为指定区域添加边框"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in range(start_row, end_row + 1):
            for col in cols:
                ws[f"{col}{row}"].border = thin_border

    def _setup_print_settings(self, ws: Worksheet, total_row_idx: int):
        """设置页面打印相关配置"""
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2
        ws.page_setup.scale = 90
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        ws.print_area = f"A1:M{total_row_idx}"

        # 【修复】打印标题行的正确设置方式，不带美元符号
        ws.print_title_rows = '48:48'

    def _adjust_row_heights_and_column_widths(self, ws: Worksheet, data_length: int):
        """调整行高和列宽以适应内容"""
        column_widths = {
            'A': 5, 'B': 18, 'C': 18, 'D': 6, 'E': 10, 'F': 9, 'G': 9,
            'H': 9, 'I': 9, 'J': 10, 'K': 10, 'L': 22, 'M': 8
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            ws.column_dimensions[col].hidden = False  # 确保列可见

        # 调整数据区域和合计行的行高
        start_row = 49  # 数据开始行
        for row in range(start_row, start_row + data_length):
            ws.row_dimensions[row].height = 30

        total_row_idx = start_row + data_length
        ws.row_dimensions[total_row_idx].height = 25

    def _clear_data_region_merges(self, ws: Worksheet, data_start_row: int = 49):
        """清理模板中数据区域的合并单元格，避免导出的Excel被修复"""
        ranges_to_clear = []
        for merge_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            if min_row >= data_start_row:
                ranges_to_clear.append(str(merge_range))

        for merge_range in ranges_to_clear:
            ws.unmerge_cells(merge_range)

    def _populate_sheet(self, ws: Worksheet, data: List[Dict], year_month: str,
                        platform_company: Optional[str] = None,
                        credit_code: Optional[str] = None):
        """用给定数据填充单个工作表"""
        if not data:
            return

        # 清理模板在数据区域遗留的合并，避免不同Excel软件解析不一致
        self._clear_data_region_merges(ws)

        company = platform_company or self.DEFAULT_COMPANY
        code = credit_code or self.DEFAULT_CREDIT_CODE
        today_cn = datetime.datetime.now().strftime('%Y年%m月%d日')
        year_month_cn = year_month.replace('-', '年') + '月'
        enterprise_name = data[0]['enterprise_name']

        # 1. 填写静态信息
        self.safe_write_merged_cell(ws, 'D10', company, 'D10:H10', align='right')
        self.safe_write_merged_cell(ws, 'D12', code, 'D12:H12', align='right')
        self.safe_write_merged_cell(ws, 'D8', year_month_cn, 'D8:H9', align='right')
        self.safe_write_merged_cell(ws, 'D14', today_cn, 'D14:H15', align='right')
        ws['C35'].value, ws['C36'].value = company, code
        ws['C37'].value, ws['C38'].value = today_cn, year_month_cn
        ws['C44'].value = enterprise_name

        # 2. 填充个人所得税纳税清单
        start_row = 49
        total_amount = 0.0

        # 统一设置数据区域字体和对齐
        for i, record in enumerate(data, start=0):
            row_idx = start_row + i
            ws[f'A{row_idx}'].value = record['序号']
            ws[f'B{row_idx}'].value = record['纳税人姓名']
            ws[f'C{row_idx}'].value = record['身份证号']
            ws[f'E{row_idx}'].value = record['营业额_元']
            ws[f'F{row_idx}'].value = record['增值税_元']
            ws[f'G{row_idx}'].value = record['教育费附加_元']
            ws[f'H{row_idx}'].value = record['地方教育附加_元']
            ws[f'I{row_idx}'].value = record['城市维护建设费_元']
            ws[f'J{row_idx}'].value = record['个人经营所得税税率']
            ws[f'K{row_idx}'].value = record['应纳个人经营所得税_元']
            ws[f'L{row_idx}'].value = record['备注']

            # 设置格式
            ws[f'C{row_idx}'].number_format = '@'
            ws[f'E{row_idx}'].number_format = '0.00'
            ws[f'F{row_idx}'].number_format = '0.00'
            ws[f'G{row_idx}'].number_format = '0.00'
            ws[f'H{row_idx}'].number_format = '0.00'
            ws[f'I{row_idx}'].number_format = '0.00'
            ws[f'J{row_idx}'].number_format = '0.00%'
            ws[f'K{row_idx}'].number_format = '0.00'

            # 统一应用字体和对齐
            for col in "ABCDEFFGHIJKLM":
                cell = ws[f'{col}{row_idx}']
                cell.font = self.DATA_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 恢复身份证号 (C/D) 与备注 (L/M) 列的合并展示
            id_merge_range = f'C{row_idx}:D{row_idx}'
            remark_merge_range = f'L{row_idx}:M{row_idx}'
            for merge_range in (id_merge_range, remark_merge_range):
                if self.is_merged(ws, merge_range):
                    ws.unmerge_cells(merge_range)
            ws.merge_cells(id_merge_range)
            ws.merge_cells(remark_merge_range)

            total_amount += float(record['营业额_元'])

        # 3. 写入合计行
        total_row_idx = start_row + len(data)
        amount_cn = self.amount_to_cn(total_amount)
        formatted_total = self.format_currency(total_amount)

        self.safe_write_merged_cell(ws, f'A{total_row_idx}', '合计：', f'A{total_row_idx}:D{total_row_idx}', bold=True)
        self.safe_write_merged_cell(ws, f'E{total_row_idx}', formatted_total, f'E{total_row_idx}:F{total_row_idx}',
                                    bold=True)
        self.safe_write_merged_cell(ws, f'G{total_row_idx}', amount_cn, f'G{total_row_idx}:M{total_row_idx}', bold=True)
        self.safe_write_merged_cell(ws, 'J45', amount_cn, 'J45:M45', bold=True)

        # 4. 添加边框
        self.add_border(ws, start_row - 1, total_row_idx,
                        ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'])

        # 5. 调整布局和打印设置
        self._adjust_row_heights_and_column_widths(ws, len(data))
        self._setup_print_settings(ws, total_row_idx)

    def generate_tax_report(self, year_month: str, output_path: Union[str, Path],
                            enterprise_ids: Optional[Union[int, List[int], str]] = None,
                            amount_type: int = 1,
                            platform_company: Optional[str] = None,
                            credit_code: Optional[str] = None):
        """
        生成税务报表
        - output_path: 输出文件路径
        - enterprise_ids: 企业ID或ID列表
        - amount_type: 金额类型 (1=含服务费, 2=不含服务费, 3=账单金额)
        - platform_company: 平台企业名称（可选，不传则使用默认值）
        - credit_code: 社会统一信用代码（可选，不传则使用默认值）
        """
        output_path = Path(output_path) if isinstance(output_path, str) else output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        all_data = self.query_tax_data(year_month, enterprise_ids, amount_type)
        if not all_data:
            logger.warning("没有查询到数据，无法生成报表")
            raise ValueError("没有查询到数据，无法生成报表")

        try:
            # 【优化】统一处理逻辑，不再区分单个或多个企业
            grouped_data = defaultdict(list)
            for record in all_data:
                grouped_data[record['enterprise_name']].append(record)

            output_wb = openpyxl.load_workbook(self.template_path)
            template_ws = output_wb.active
            template_sheet_name = template_ws.title

            for enterprise_name, records in grouped_data.items():
                ws = output_wb.copy_worksheet(template_ws)
                # 确保工作表名称合法
                safe_name = "".join(c for c in enterprise_name if c not in r'[]:*?/\ ')
                ws.title = safe_name[:31]  # Excel工作表名称最大长度为31字符

                for i, record in enumerate(records, start=1):
                    record['序号'] = i

                self._populate_sheet(ws, records, year_month, platform_company, credit_code)

            # 所有企业复制完成后再删除模板，避免复制过程中缺失基础格式
            if template_sheet_name in output_wb.sheetnames:
                del output_wb[template_sheet_name]

            output_wb.save(output_path)
            logger.info(f"报表已生成: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"生成报表时出错: {str(e)}", exc_info=True)
            raise


class PlatformReportGenerator:
    """平台内经营者和从业人员报送表生成器"""

    # 模板路径配置
    INCOME_TEMPLATE_NAME = "平台内的经营者和从业人员收入信息报送表01.xlsx"
    IDENTITY_TEMPLATE_NAME = "平台内的经营者和从业人员身份信息报送表01.xlsx"

    # 默认值配置
    DEFAULT_COMPANY = "春苗人力资源（北京）有限公司"
    DEFAULT_CREDIT_CODE = "91110106MACE710372"
    DEFAULT_PLATFORM_NAME = "灵零猴交易合规数字化服务平台"
    DEFAULT_FONT = Font(name='宋体', size=10)
    DATA_FONT = Font(name='宋体', size=9)
    BOLD_FONT = Font(name='宋体', size=10, bold=True)

    def __init__(self, db_config: dict):
        self.db_config = db_config
        self.template_dir = self._get_template_dir()

    def _get_template_dir(self) -> Path:
        """获取模板目录的跨平台路径"""
        app_dir = Path(__file__).parent
        template_dir = app_dir / "templates"
        if not template_dir.exists():
            raise FileNotFoundError(f"模板目录未找到: {template_dir}")
        return template_dir

    def _get_template_path(self, template_name: str) -> Path:
        """获取模板文件的跨平台路径，支持模糊匹配（用于匹配带数字后缀的模板如01/02/03）"""
        # 如果精确匹配的文件存在，直接返回
        exact_path = self.template_dir / template_name
        if exact_path.exists():
            return exact_path

        # 提取文件名模式（不含数字后缀部分），用于模糊匹配
        # 例如 "平台内的经营者和从业人员收入信息报送表*.xlsx"
        # 将 "平台内的经营者和从业人员收入信息报送表01.xlsx" 转为 "平台内的经营者和从业人员收入信息报送表*.xlsx"
        import re
        # 匹配末尾的数字部分（01, 02, 03等）
        pattern = re.sub(r'\d+(\.[^.]+)$', r'*\1', template_name)

        # 在模板目录中查找匹配的文件
        matching_files = list(self.template_dir.glob(pattern))
        if matching_files:
            # 返回第一个匹配的文件
            return matching_files[0]

        raise FileNotFoundError(f"Excel模板文件未找到: {template_name} (已尝试模糊匹配模式: {pattern})")

    def _format_date_range(self, start_date: str, end_date: str) -> str:
        """格式化日期范围"""
        start = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        return f"{start.strftime('%Y年%m月%d日')}至{end.strftime('%Y年%m月%d日')}"

    def _format_date(self, date_str: str = None) -> str:
        """格式化日期为 年 月 日"""
        if date_str is None:
            return datetime.datetime.now().strftime('%Y年%m月%d日')
        try:
            if isinstance(date_str, datetime.datetime):
                return date_str.strftime('%Y年%m月%d日')
            dt = datetime.datetime.strptime(str(date_str)[:10], '%Y-%m-%d')
            return dt.strftime('%Y年%m月%d日')
        except Exception:
            return str(date_str) if date_str else ''

    def query_platform_data(self, start_date: str, end_date: str,
                            enterprise_ids: Optional[Union[int, List[int], str]] = None,
                            amount_type: int = 1,
                            tax_id: Optional[int] = None) -> List[Dict]:
        """
        查询平台内经营者和从业人员数据
        amount_type: 1=含服务费(pay_amount), 2=不含服务费(worker_pay_amount), 3=账单金额(bill_amount)
        tax_id: 运营主体ID，不传或传0表示查全部
        """
        # 根据金额类型选择金额字段
        if amount_type == 1:
            amount_field = "w.pay_amount"
        elif amount_type == 2:
            amount_field = "w.worker_pay_amount"
        else:
            amount_field = "w.bill_amount"

        try:
            with DatabaseManager(self.db_config) as conn:
                with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = f"""
                    SELECT
                        w.realname AS name,
                        w.credential_num AS credential_num,
                        MAX(w.payment_account) AS payment_account,
                        eb.enterprise_name AS enterprise_name,
                        eb.USCC AS uscc,
                        ROUND(SUM({amount_field}), 2) AS labor_income,
                        ROUND(SUM(w.service_amount), 2) AS service_fee,
                        COUNT(w.id) AS trade_count,
                        MAX(wa.username) AS miniapp_id,
                        MAX(w.mobile) AS mobile,
                        ROUND(SUM(w.tax_amount), 2) AS tax_amount,
                        ROUND(SUM(w.worker_pay_amount + w.tax_amount), 2) AS total_with_tax,
                        COALESCE(MIN(es.sign_time), MIN(w.payment_over_time)) AS sign_time
                    FROM
                        biz_balance_worker w
                        LEFT JOIN biz_enterprise_base eb ON eb.id = w.enterprise_id
                        LEFT JOIN biz_worker_account wa ON wa.worker_id = w.worker_id
                        LEFT JOIN (
                            SELECT worker_id, MIN(sign_time) AS sign_time
                            FROM biz_enterprise_sign
                            WHERE sign_status = 0
                            GROUP BY worker_id
                        ) es ON es.worker_id = w.worker_id
                    WHERE
                        w.pay_status = 3
                        AND w.enterprise_id <> 36
                        AND w.payment_over_time >= %s
                        AND w.payment_over_time < %s
                    """
                    params = [start_date, end_date]

                    if enterprise_ids is not None:
                        ids_to_query = []
                        if isinstance(enterprise_ids, int):
                            ids_to_query = [enterprise_ids]
                        elif isinstance(enterprise_ids, str):
                            ids_to_query = [int(eid.strip()) for eid in enterprise_ids.split(',') if eid.strip()]
                        elif isinstance(enterprise_ids, list):
                            ids_to_query = enterprise_ids

                        if ids_to_query:
                            placeholders = ', '.join(['%s'] * len(ids_to_query))
                            sql += f" AND w.enterprise_id IN ({placeholders})"
                            params.extend(ids_to_query)

                    if tax_id:
                        sql += f" AND w.tax_id = %s"
                        params.append(tax_id)

                    sql += """
                    GROUP BY
                        w.worker_id,
                        w.realname,
                        w.credential_num,
                        eb.enterprise_name,
                        eb.USCC
                    ORDER BY
                        w.realname, eb.enterprise_name
                    """

                    # ===== 调试打印结束 =====

                    cursor.execute(sql, tuple(params))
                    rows = cursor.fetchall()
                    # 转换 Decimal 为 float 以支持 JSON 序列化
                    for row in rows:
                        for key, value in row.items():
                            if isinstance(value, Decimal):
                                row[key] = float(value)
                    return rows
        except Exception as e:
            logger.error(f"数据库查询错误: {str(e)}")
            return []

    def is_merged(self, ws: Worksheet, merge_range: str) -> bool:
        """检查单元格范围是否已合并"""
        return merge_range in [str(r) for r in ws.merged_cells.ranges]

    def _populate_income_sheet(self, ws: Worksheet, data: List[Dict],
                               platform_company: str, platform_name: str,
                               credit_code: str, start_date: str, end_date: str):
        """填充收入信息表"""
        if not data:
            return

        # 1. 填写表头信息 (行3-4)
        date_range = self._format_date_range(start_date, end_date)
        today_cn = self._format_date()

        # 居左不换行的对齐方式
        left_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # 行3: 企业名称和平台名称（不换行，居左）
        # A3保持不变，企业名称插入在D3
        ws['D3'].value = platform_company
        ws['D3'].alignment = left_no_wrap

        ws['P3'].value = f"平台名称：{platform_name}"
        ws['P3'].alignment = left_no_wrap

        # 行4: 信用代码和日期范围（不换行，居左）
        ws['A4'].value = f"互联网平台企业统一社会信用代码（纳税人识别号）：{credit_code}"
        ws['A4'].alignment = left_no_wrap

        ws['P4'].value = f"收入所属期间：{date_range}"
        ws['P4'].alignment = left_no_wrap

        ws['AD4'].value = f"报送日期：{today_cn}"
        ws['AD4'].alignment = left_no_wrap

        # 2. 动态插入数据行 - 保护底部内容
        # 定位声明行（保护声明行及以下内容）
        declaration_row = self._find_declaration_row(ws, "谨声明")
        if declaration_row == -1:
            declaration_row = 20  # 默认保护行

        # 数据从第9行开始（保持第8行列标题不动）
        data_start_row = 9
        n_new = len(data)

        # 先取消数据区域的合并单元格（避免写入时出错）
        # 取消所有与数据写入区域重叠的合并单元格
        ranges_to_unmerge = []
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            # 如果合并区域与数据区域(第data_start_row行到declaration_row-1行)有交集
            if max_row > data_start_row and min_row < declaration_row:
                ranges_to_unmerge.append(str(merge_range))

        for merge_range in ranges_to_unmerge:
            ws.unmerge_cells(merge_range)

        # 如果需要更多行，在声明行前插入空行
        if n_new > 0:
            ws.insert_rows(declaration_row, n_new)

        # 为所有待写入的数据行应用格式（从第10行模板行复制）
        # 记录模板行的合并单元格信息，用于后续重新合并
        template_merges = []
        for merge_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            if min_row <= data_start_row <= max_row:
                template_merges.append(str(merge_range))

        for i in range(n_new):
            target_row = data_start_row + i
            # 复制第10行格式到新插入的行
            for col in range(1, 34):
                src_cell = ws.cell(row=data_start_row, column=col)
                tgt_cell = ws.cell(row=target_row, column=col)
                self._copy_cell_style(src_cell, tgt_cell)
            # 复制行高
            ws.row_dimensions[target_row].height = ws.row_dimensions[data_start_row].height

            # 取消目标行中所有与模板行相同的合并单元格
            # 确保写入数据时不会遇到合并单元格
            ranges_to_unmerge_on_target = []
            for merge_range in list(ws.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
                # 如果合并范围跨越目标行
                if min_row <= target_row <= max_row:
                    ranges_to_unmerge_on_target.append(str(merge_range))
            for merge_range in ranges_to_unmerge_on_target:
                try:
                    ws.unmerge_cells(merge_range)
                except Exception:
                    pass

        # 写入数据到各列
        for i, record in enumerate(data):
            row_idx = data_start_row + i
            # 序号
            ws.cell(row=row_idx, column=1, value=i + 1)
            # 是否已取得登记证照 = 否
            ws.cell(row=row_idx, column=2, value="否")
            # 未取得登记证照 - 姓名
            ws.cell(row=row_idx, column=5, value=record.get('name', ''))
            # 证件类型
            ws.cell(row=row_idx, column=6, value="201|居民身份证")
            # 证件号码
            ws.cell(row=row_idx, column=7, value=record.get('credential_num', ''))
            # 国家或地区
            ws.cell(row=row_idx, column=8, value="156|中国")
            # 收入来源的互联网平台名称
            ws.cell(row=row_idx, column=9, value="灵零猴交易合规数字化服务平台")
            # 收入来源的店铺名称
            ws.cell(row=row_idx, column=10, value=record.get('enterprise_name', ''))
            # 收入来源的店铺唯一标识码 = 企业信用代码
            ws.cell(row=row_idx, column=11, value=record.get('uscc', ''))

            # 收入总额 (U列=21) = 劳务报酬 + 服务费
            labor_income = float(record.get('labor_income', 0) or 0)
            service_fee = float(record.get('service_fee', 0) or 0)
            ws.cell(row=row_idx, column=21, value=labor_income + service_fee)

            # 劳务报酬 (Z列=26)
            ws.cell(row=row_idx, column=26, value=labor_income)

            # 从事其他网络交易活动取得的收入（净额）- AC列=29 = 默认0
            ws.cell(row=row_idx, column=29, value=0)

            # 服务费合计 (AD列=30)
            ws.cell(row=row_idx, column=30, value=service_fee)

            # 交易笔数 (AE列=31)
            trade_count = int(record.get('trade_count', 0) or 0)
            ws.cell(row=row_idx, column=31, value=trade_count)

            # 个税 (AF列=32)
            tax_amount = float(record.get('tax_amount', 0) or 0)
            ws.cell(row=row_idx, column=32, value=tax_amount)

            # 含税金额 (AG列=33) = worker_pay_amount + tax_amount
            total_with_tax = float(record.get('total_with_tax', 0) or 0)
            ws.cell(row=row_idx, column=33, value=total_with_tax)

        # 设置所有数据单元格居中对齐
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in range(n_new):
            target_row = data_start_row + i
            for col in range(1, 34):
                cell = ws.cell(row=target_row, column=col)
                cell.alignment = center_alignment

        # 重新应用合并单元格（每行单独合并）
        for i in range(n_new):
            target_row = data_start_row + i
            for merge_range_str in template_merges:
                min_col, min_row, max_col, max_row = range_boundaries(merge_range_str)
                # 计算新的合并范围：保持列不变，行调整为当前行
                new_min_row = target_row
                new_max_row = target_row
                col_letter_start = get_column_letter(min_col)
                col_letter_end = get_column_letter(max_col)
                new_merge_range = f"{col_letter_start}{new_min_row}:{col_letter_end}{new_max_row}"
                try:
                    ws.merge_cells(new_merge_range)
                except Exception:
                    pass

        # 添加声明、签章和经办人信息（收入信息表）
        self._add_declaration_and_signature(ws, data_start_row, n_new, is_income_table=True)

    def _add_declaration_and_signature(self, ws: Worksheet, data_start_row: int, n_new: int,
                                       is_income_table: bool = False):
        """在数据区域下方添加声明、签章和经办人信息

        Args:
            is_income_table: True=收入信息表, False=身份信息表
        """
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(name='宋体', size=10, bold=True)

        # 根据表格类型确定合并列和签章列
        if is_income_table:
            # 收入信息表：合并到 AE 列，签章在 O 列，经办人合并到 N 列，受理人合并到 AE 列
            last_col_letter = 'AE'
            signature_col = 15  # O列
            agent_end_col_letter = 'N'
            agent_end_col = 14  # N列
            recipient_end_col_letter = 'AE'
            recipient_end_col = 31  # AE列
        else:
            # 身份信息表：合并到 W 列，签章在 N 列，经办人合并到 M 列，受理人合并到 W 列
            last_col_letter = 'W'
            signature_col = 14  # N列
            agent_end_col_letter = 'M'
            agent_end_col = 13  # M列
            recipient_end_col_letter = 'W'
            recipient_end_col = 23  # W列

        # 第1行：声明行（在数据最后一行后面）
        statement_row = data_start_row + n_new
        ws.row_dimensions[statement_row].height = 60
        ws.merge_cells(f'A{statement_row}:{last_col_letter}{statement_row}')

        if is_income_table:
            statement_text = ('谨声明：本表是根据国家有关法律、行政法规及相关规定填报的，是真实的、准确的、完整的。'
                             '平台内的经营者和从业人员的当期收入信息均已如实报送；除免于报送收入信息的情形外，'
                             '其他未报送收入信息的平台内的经营者和从业人员在当期均未发生网络交易活动，当期收入总额、退款金额均为零。')
        else:
            statement_text = '谨声明：本表是根据国家有关法律、行政法规及相关规定填报的，是真实的、准确的、完整的。'

        ws.cell(row=statement_row, column=1, value=statement_text)
        ws.cell(row=statement_row, column=1).alignment = center_alignment
        ws.cell(row=statement_row, column=1).font = bold_font

        # 第2行：互联网平台企业（签章）：年  月  日
        signature_row = statement_row + 1
        ws.row_dimensions[signature_row].height = 25
        ws.cell(row=signature_row, column=signature_col, value='互联网平台企业（签章）：              年   月   日')
        ws.cell(row=signature_row, column=signature_col).alignment = left_alignment

        # 第3行：经办人 + 受理人（受理人包含：受理人、受理税务机关（章）、受理时间），行高90
        agent_row = signature_row + 1
        ws.row_dimensions[agent_row].height = 90

        # 经办人： + 空行 + 经办人联系方式：（A到agent_end_col合并，换行显示）
        ws.merge_cells(f'A{agent_row}:{agent_end_col_letter}{agent_row}')
        ws.cell(row=agent_row, column=1, value='经办人：\n\n经办人联系方式：')
        ws.cell(row=agent_row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 受理人（从agent_end_col+1到recipient_end_col合并，包含受理人、受理税务机关（章）、受理时间）
        recipient_start_col = agent_end_col + 1
        recipient_start_letter = get_column_letter(recipient_start_col)
        ws.merge_cells(f'{recipient_start_letter}{agent_row}:{recipient_end_col_letter}{agent_row}')
        ws.cell(row=agent_row, column=recipient_start_col,
                value='受理人：\n\n受理税务机关（章）：\n\n受理时间：     年   月   日')
        ws.cell(row=agent_row, column=recipient_start_col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _populate_identity_sheet(self, ws: Worksheet, data: List[Dict],
                                  platform_company: str, platform_name: str,
                                  credit_code: str):
        """填充身份信息表"""
        if not data:
            return

        # 居左不换行的对齐方式
        left_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # 1. 填写表头信息
        today_cn = self._format_date()

        # A3: 互联网平台企业名称（全量覆盖）
        ws['A3'].value = f"互联网平台企业名称：{platform_company}"
        ws['A3'].alignment = left_no_wrap

        # F3: 互联网平台企业统一社会信用代码（纳税人识别号）：信用代码
        ws['F3'].value = f"互联网平台企业统一社会信用代码（纳税人识别号）：{credit_code}"
        ws['F3'].alignment = left_no_wrap

        # W3: 报送日期 - 取消合并后写入
        for merge_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            if min_row == 3 and min_col <= 23 <= max_col:  # W列是第23列
                ws.unmerge_cells(str(merge_range))

        # Q3: 平台名称：+ 平台名称
        ws['Q3'].value = f"平台名称：{platform_name}"
        ws['Q3'].alignment = left_no_wrap

        # W3: 报送日期：   年  月  日（居右）
        ws['W3'].value = f"报送日期：{today_cn}"
        ws['W3'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

        # 2. 动态插入数据行 - 保护声明行
        # 定位声明行
        declaration_row = self._find_declaration_row(ws, "谨声明")
        if declaration_row == -1:
            declaration_row = 20  # 默认保护行

        # 数据从第6行开始（保持第5行列标题不动）
        data_start_row = 6
        n_new = len(data)

        # 先取消数据区域的合并单元格（避免写入时出错）
        # 取消所有与数据写入区域重叠的合并单元格
        # 记录模板行的合并单元格信息，用于后续重新合并
        template_merges = []
        for merge_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            # 如果合并区域与数据区域(第data_start_row行到declaration_row-1行)有交集
            if max_row > data_start_row and min_row < declaration_row:
                template_merges.append(str(merge_range))

        for merge_range in template_merges:
            ws.unmerge_cells(merge_range)

        # 如果需要更多行，在声明行前插入空行
        if n_new > 0:
            ws.insert_rows(declaration_row, n_new)

        # 为所有待写入的数据行应用格式（从第7行模板行复制）
        for i in range(n_new):
            target_row = data_start_row + i
            for col in range(1, 24):
                src_cell = ws.cell(row=data_start_row, column=col)
                tgt_cell = ws.cell(row=target_row, column=col)
                self._copy_cell_style(src_cell, tgt_cell)
            # 复制行高
            ws.row_dimensions[target_row].height = ws.row_dimensions[data_start_row].height

            # 取消目标行中所有与模板行相同的合并单元格
            # 确保写入数据时不会遇到合并单元格
            ranges_to_unmerge_on_target = []
            for merge_range in list(ws.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
                # 如果合并范围跨越目标行
                if min_row <= target_row <= max_row:
                    ranges_to_unmerge_on_target.append(str(merge_range))
            for merge_range in ranges_to_unmerge_on_target:
                try:
                    ws.unmerge_cells(merge_range)
                except Exception:
                    pass

        # 写入数据到各列
        for i, record in enumerate(data):
            row_idx = data_start_row + i
            # 序号
            ws.cell(row=row_idx, column=1, value=i + 1)
            # 是否已取得登记证照 = 否
            ws.cell(row=row_idx, column=2, value="否")
            # 未取得登记证照 - 姓名
            ws.cell(row=row_idx, column=6, value=record.get('name', ''))
            # 证件类型
            ws.cell(row=row_idx, column=7, value="201|居民身份证")
            # 证件号码
            ws.cell(row=row_idx, column=8, value=record.get('credential_num', ''))
            # 国家或地区
            ws.cell(row=row_idx, column=9, value="156|中国")
            # 是否存在免于报送收入信息情形 = 否
            ws.cell(row=row_idx, column=10, value="否")
            # 店铺（用户）名称
            ws.cell(row=row_idx, column=13, value=record.get('enterprise_name', ''))
            # 店铺（用户）唯一标识码
            ws.cell(row=row_idx, column=14, value=record.get('uscc', ''))
            # 账户名称 (列17)
            ws.cell(row=row_idx, column=17, value=record.get('name', ''))
            # 银行账号/支付账户 (列18)
            ws.cell(row=row_idx, column=18, value=record.get('payment_account', ''))
            # 联系人姓名
            ws.cell(row=row_idx, column=19, value=record.get('name', ''))
            # 联系电话
            ws.cell(row=row_idx, column=20, value=record.get('mobile', ''))
            # 经营开始时间 (格式: yyyy-MM-dd)
            经营时间 = record.get('sign_time')
            if 经营时间:
                try:
                    if isinstance(经营时间, datetime.datetime):
                        ws.cell(row=row_idx, column=21, value=经营时间.strftime('%Y-%m-%d'))
                    else:
                        dt = datetime.datetime.strptime(str(经营时间)[:10], '%Y-%m-%d')
                        ws.cell(row=row_idx, column=21, value=dt.strftime('%Y-%m-%d'))
                except Exception:
                    ws.cell(row=row_idx, column=21, value=str(经营时间) if 经营时间 else '')
            # 信息状态标识 = 新增
            ws.cell(row=row_idx, column=23, value="新增")

        # 设置所有数据单元格居中对齐
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in range(n_new):
            target_row = data_start_row + i
            for col in range(1, 24):
                cell = ws.cell(row=target_row, column=col)
                cell.alignment = center_alignment

        # 重新应用合并单元格（每行单独合并）
        for i in range(n_new):
            target_row = data_start_row + i
            for merge_range_str in template_merges:
                min_col, min_row, max_col, max_row = range_boundaries(merge_range_str)
                # 计算新的合并范围：保持列不变，行调整为当前行
                new_min_row = target_row
                new_max_row = target_row
                col_letter_start = get_column_letter(min_col)
                col_letter_end = get_column_letter(max_col)
                new_merge_range = f"{col_letter_start}{new_min_row}:{col_letter_end}{new_max_row}"
                try:
                    ws.merge_cells(new_merge_range)
                except Exception:
                    pass

        # 添加声明、签章和经办人信息
        self._add_declaration_and_signature(ws, data_start_row, n_new, is_income_table=False)

    def _clear_data_region_merges(self, ws: Worksheet, data_start_row: int):
        """清理模板中数据区域的合并单元格，避免导出的Excel被修复"""
        ranges_to_clear = []
        for merge_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            if min_row >= data_start_row:
                ranges_to_clear.append(str(merge_range))

        for merge_range in ranges_to_clear:
            ws.unmerge_cells(merge_range)

    def _write_cell(self, ws: Worksheet, cell_ref: str, value, font: Font = None):
        """写入单元格"""
        cell = ws[cell_ref]
        cell.value = value
        if font:
            cell.font = font
        else:
            cell.font = self.DEFAULT_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _copy_cell_style(self, src_cell, tgt_cell):
        """复制单元格样式"""
        if src_cell.has_style:
            tgt_cell.font = src_cell.font.copy()
            tgt_cell.border = src_cell.border.copy()
            tgt_cell.fill = src_cell.fill.copy()
            tgt_cell.number_format = src_cell.number_format
            tgt_cell.protection = src_cell.protection.copy()
            tgt_cell.alignment = src_cell.alignment.copy()

    def _find_declaration_row(self, ws: Worksheet, keyword: str = "谨声明") -> int:
        """查找包含关键字的行号，返回-1如果未找到"""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and keyword in str(cell.value):
                    return cell.row
        return -1

    def generate_income_report(self, output_path: Union[str, Path],
                               start_date: str, end_date: str,
                               platform_company: str = None,
                               platform_name: str = None,
                               credit_code: str = None,
                               enterprise_ids: Optional[Union[int, List[int], str]] = None,
                               amount_type: int = 1,
                               tax_id: Optional[int] = None):
        """
        生成平台收入信息表
        - output_path: 输出文件路径
        - start_date: 开始日期 YYYY-MM-DD
        - end_date: 结束日期 YYYY-MM-DD
        - platform_company: 平台企业名称
        - platform_name: 平台名称
        - credit_code: 统一社会信用代码
        - enterprise_ids: 企业ID筛选
        - amount_type: 金额类型
        - tax_id: 运营主体ID，不传或传0表示查全部
        """
        output_path = Path(output_path) if isinstance(output_path, str) else output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        company = platform_company or self.DEFAULT_COMPANY
        platform = platform_name or self.DEFAULT_PLATFORM_NAME
        code = credit_code or self.DEFAULT_CREDIT_CODE

        all_data = self.query_platform_data(start_date, end_date, enterprise_ids, amount_type, tax_id)
        if not all_data:
            logger.warning("没有查询到数据，无法生成报表")
            raise ValueError("没有查询到数据，无法生成报表")

        try:
            # 加载收入信息表模板
            income_wb = openpyxl.load_workbook(self._get_template_path(self.INCOME_TEMPLATE_NAME))
            income_ws = income_wb.active

            # 填充收入信息表数据（动态插入行，保护声明行）
            self._populate_income_sheet(income_ws, all_data, company, platform, code, start_date, end_date)

            income_wb.save(output_path)
            logger.info(f"收入信息表已生成: {output_path}")
            return str(output_path), len(all_data)

        except Exception as e:
            logger.error(f"生成收入信息表时出错: {str(e)}", exc_info=True)
            raise

    def generate_identity_report(self, output_path: Union[str, Path],
                                start_date: str, end_date: str,
                                platform_company: str = None,
                                platform_name: str = None,
                                credit_code: str = None,
                                enterprise_ids: Optional[Union[int, List[int], str]] = None,
                                amount_type: int = 1,
                                tax_id: Optional[int] = None):
        """
        生成平台身份信息表
        - output_path: 输出文件路径
        - start_date: 开始日期 YYYY-MM-DD
        - end_date: 结束日期 YYYY-MM-DD
        - platform_company: 平台企业名称
        - platform_name: 平台名称
        - credit_code: 统一社会信用代码
        - enterprise_ids: 企业ID筛选
        - amount_type: 金额类型
        - tax_id: 运营主体ID，不传或传0表示查全部
        """
        output_path = Path(output_path) if isinstance(output_path, str) else output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        company = platform_company or self.DEFAULT_COMPANY
        platform = platform_name or self.DEFAULT_PLATFORM_NAME
        code = credit_code or self.DEFAULT_CREDIT_CODE

        all_data = self.query_platform_data(start_date, end_date, enterprise_ids, amount_type, tax_id)
        if not all_data:
            logger.warning("没有查询到数据，无法生成报表")
            raise ValueError("没有查询到数据，无法生成报表")

        try:
            # 加载身份信息表模板
            identity_wb = openpyxl.load_workbook(self._get_template_path(self.IDENTITY_TEMPLATE_NAME))
            identity_ws = identity_wb.active

            # 填充身份信息表数据（动态插入行，保护声明行）
            self._populate_identity_sheet(identity_ws, all_data, company, platform, code)

            identity_wb.save(output_path)
            logger.info(f"身份信息表已生成: {output_path}")
            return str(output_path), len(all_data)

        except Exception as e:
            logger.error(f"生成身份信息表时出错: {str(e)}", exc_info=True)
            raise

    def generate_combined_report(self, output_path: Union[str, Path],
                                 start_date: str, end_date: str,
                                 platform_company: str = None,
                                 platform_name: str = None,
                                 credit_code: str = None,
                                 enterprise_ids: Optional[Union[int, List[int], str]] = None,
                                 amount_type: int = 1,
                                 tax_id: Optional[int] = None):
        """
        生成组合报表（收入信息表 + 身份信息表在同一个Excel的不同sheet中）
        用于手动复制数据到模板
        """
        output_path = Path(output_path) if isinstance(output_path, str) else output_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        all_data = self.query_platform_data(start_date, end_date, enterprise_ids, amount_type, tax_id)
        if not all_data:
            logger.warning("没有查询到数据，无法生成报表")
            raise ValueError("没有查询到数据，无法生成报表")

        try:
            wb = openpyxl.Workbook()

            # 创建收入信息表
            income_ws = wb.active
            income_ws.title = "收入信息表"
            self._populate_combined_income_sheet(income_ws, all_data, platform_company, platform_name, credit_code, start_date, end_date)

            # 创建身份信息表
            identity_ws = wb.create_sheet(title="身份信息表")
            self._populate_combined_identity_sheet(identity_ws, all_data, platform_company, platform_name, credit_code)

            wb.save(output_path)
            logger.info(f"组合报表已生成: {output_path}")
            return str(output_path), len(all_data)

        except Exception as e:
            logger.error(f"生成组合报表时出错: {str(e)}", exc_info=True)
            raise

    def _populate_combined_income_sheet(self, ws: Worksheet, data: List[Dict],
                                        platform_company: str, platform_name: str,
                                        credit_code: str, start_date: str, end_date: str):
        """填充组合报表的收入信息表"""
        if not data:
            return

        # 写表头（按模板列顺序）
        # 序号=1, 是否取得登记证照=2, 未取得登记证照-姓名=5, 证件类型=6, 证件号码=7,
        # 国家或地区=8, 收入来源互联网平台名称=9, 店铺名称=10, 店铺唯一标识码=11,
        # 收入总额=U(21), 劳务报酬=Z(26), AC(29)=从事其他网络交易收入(默认0),
        # AD(30)=服务费总计, AE(31)=笔数

        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="是否已取得登记证照")
        ws.cell(row=1, column=5, value="未取得登记证照-姓名")
        ws.cell(row=1, column=6, value="证件类型")
        ws.cell(row=1, column=7, value="证件号码")
        ws.cell(row=1, column=8, value="国家或地区")
        ws.cell(row=1, column=9, value="收入来源的互联网平台名称")
        ws.cell(row=1, column=10, value="收入来源的店铺名称")
        ws.cell(row=1, column=11, value="收入来源的店铺唯一标识码")
        ws.cell(row=1, column=21, value="收入总额")
        ws.cell(row=1, column=26, value="劳务报酬")
        ws.cell(row=1, column=29, value="从事其他网络交易活动取得的收入（净额）")
        ws.cell(row=1, column=30, value="服务费合计")
        ws.cell(row=1, column=31, value="交易笔数")
        ws.cell(row=1, column=32, value="个税")
        ws.cell(row=1, column=33, value="含税金额")

        # 填充数据
        for i, record in enumerate(data, start=1):
            row_idx = i + 1  # 数据从第2行开始
            labor_income = float(record.get('labor_income', 0) or 0)
            service_fee = float(record.get('service_fee', 0) or 0)
            trade_count = int(record.get('trade_count', 0) or 0)

            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value="否")
            ws.cell(row=row_idx, column=5, value=record.get('name', ''))
            ws.cell(row=row_idx, column=6, value="201|居民身份证")
            ws.cell(row=row_idx, column=7, value=record.get('credential_num', ''))
            ws.cell(row=row_idx, column=8, value="156|中国")
            ws.cell(row=row_idx, column=9, value="灵零猴交易合规数字化服务平台")
            ws.cell(row=row_idx, column=10, value=record.get('enterprise_name', ''))
            ws.cell(row=row_idx, column=11, value=record.get('uscc', ''))
            # 收入总额(U列) = 劳务报酬 + 服务费
            ws.cell(row=row_idx, column=21, value=labor_income + service_fee)
            # 劳务报酬(Z列)
            ws.cell(row=row_idx, column=26, value=labor_income)
            # 从事其他网络交易活动取得的收入（净额）(AC列) = 默认0
            ws.cell(row=row_idx, column=29, value=0)
            # 服务费合计(AD列)
            ws.cell(row=row_idx, column=30, value=service_fee)
            # 交易笔数(AE列)
            ws.cell(row=row_idx, column=31, value=trade_count)
            # 个税(AF列)
            tax_amount = float(record.get('tax_amount', 0) or 0)
            ws.cell(row=row_idx, column=32, value=tax_amount)
            # 含税金额(AG列)
            total_with_tax = float(record.get('total_with_tax', 0) or 0)
            ws.cell(row=row_idx, column=33, value=total_with_tax)

            # 设置数据字体和对齐
            for col in range(1, 33):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self.DATA_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 设置行高
            ws.row_dimensions[row_idx].height = 30

    def _populate_combined_identity_sheet(self, ws: Worksheet, data: List[Dict],
                                         platform_company: str, platform_name: str,
                                         credit_code: str):
        """填充组合报表的身份信息表"""
        if not data:
            return

        # 写表头
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="是否已取得登记证照")
        ws.cell(row=1, column=6, value="未取得登记证照-姓名")
        ws.cell(row=1, column=7, value="证件类型")
        ws.cell(row=1, column=8, value="证件号码")
        ws.cell(row=1, column=9, value="国家或地区")
        ws.cell(row=1, column=10, value="是否存在免于报送收入信息情形")
        ws.cell(row=1, column=13, value="店铺（用户）名称")
        ws.cell(row=1, column=14, value="店铺（用户）唯一标识码")
        ws.cell(row=1, column=17, value="账户名称")
        ws.cell(row=1, column=18, value="银行账号/支付账户")
        ws.cell(row=1, column=19, value="联系人姓名")
        ws.cell(row=1, column=20, value="联系电话")
        ws.cell(row=1, column=21, value="经营开始时间")
        ws.cell(row=1, column=23, value="信息状态标识")

        # 填充数据
        for i, record in enumerate(data, start=1):
            row_idx = i + 1  # 数据从第2行开始
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value="否")
            ws.cell(row=row_idx, column=6, value=record.get('name', ''))
            ws.cell(row=row_idx, column=7, value="201|居民身份证")
            ws.cell(row=row_idx, column=8, value=record.get('credential_num', ''))
            ws.cell(row=row_idx, column=9, value="156|中国")
            ws.cell(row=row_idx, column=10, value="否")
            ws.cell(row=row_idx, column=13, value=record.get('enterprise_name', ''))
            ws.cell(row=row_idx, column=14, value=record.get('uscc', ''))
            ws.cell(row=row_idx, column=17, value=record.get('name', ''))
            ws.cell(row=row_idx, column=18, value=record.get('payment_account', ''))
            ws.cell(row=row_idx, column=19, value=record.get('name', ''))
            ws.cell(row=row_idx, column=20, value=record.get('mobile', ''))
            # 经营开始时间 (格式: yyyy-MM-dd)
            经营时间 = record.get('sign_time')
            if 经营时间:
                try:
                    if isinstance(经营时间, datetime.datetime):
                        ws.cell(row=row_idx, column=21, value=经营时间.strftime('%Y-%m-%d'))
                    else:
                        dt = datetime.datetime.strptime(str(经营时间)[:10], '%Y-%m-%d')
                        ws.cell(row=row_idx, column=21, value=dt.strftime('%Y-%m-%d'))
                except Exception:
                    ws.cell(row=row_idx, column=21, value=str(经营时间) if 经营时间 else '')
            ws.cell(row=row_idx, column=23, value="新增")

            # 设置数据字体和对齐
            for col in range(1, 24):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = self.DATA_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 设置行高
            ws.row_dimensions[row_idx].height = 30
